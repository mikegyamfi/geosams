import calendar
import hashlib
import hmac
from datetime import datetime

import pandas as pd
from decouple import config
from django.contrib.auth.forms import PasswordResetForm
from django.db.models import Sum
from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponseRedirect, HttpResponse
import requests
from django.template.loader import render_to_string
from django.utils import timezone
from django.utils.http import urlsafe_base64_encode
from django.views.decorators.csrf import csrf_exempt
import json

from . import forms
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from . import helper, models
from .forms import UploadFileForm
from .models import CustomUser
from django_ratelimit.decorators import ratelimit


@login_required(login_url='login')
# Create your views here.
def home(request):
    if models.Announcement.objects.filter(active=True).exists():
        announcement = models.Announcement.objects.filter(active=True).first()
        messages.info(request, announcement.message)
        agent_price = models.AdminInfo.objects.filter().first().agent_price
        context = {
            "announcement": announcement.message if announcement else None,
            "agent_price": agent_price,
        }
        return render(request, "layouts/index.html", context=context)
    agent_price = models.AdminInfo.objects.filter().first().agent_price
    context = {
        "agent_price": agent_price,
    }
    return render(request, "layouts/index.html", context=context)


@login_required(login_url='login')
def register_as_agent(request):
    agent_price = models.AdminInfo.objects.filter().first().agent_price
    url = "https://payproxyapi.hubtel.com/items/initiate"

    reference = helper.ref_generator()

    details = {
        "amount": agent_price,
    }

    new_payment = models.Payment.objects.create(
        user=request.user,
        reference=reference,
        transaction_date=datetime.now(),
        transaction_details=details,
        channel="agent",
    )
    new_payment.save()

    payload = json.dumps({
        "totalAmount": float(agent_price) + (1 / 100) * float(agent_price),
        "description": "N/A",
        "callbackUrl": "https://www.geosams.com/hubtel_webhook",
        "returnUrl": "https://www.geosams.com",
        "cancellationUrl": "https://www.geosams.com",
        "merchantAccountNumber": "2021482",
        "clientReference": new_payment.reference
    })
    headers = {
        'Content-Type': 'application/json',
        'Authorization': config("HUBTEL_TOKEN")
    }

    response = requests.request("POST", url, headers=headers, data=payload)

    data = response.json()

    checkoutUrl = data['data']['checkoutUrl']

    return redirect(checkoutUrl)


@login_required(login_url='login')
def register_as_agent_wallet(request):
    agent_price = models.AdminInfo.objects.filter().first().agent_price
    user = models.CustomUser.objects.get(id=request.user.id)
    if user.status == "Agent" or user.status == "Super Agent":
        return JsonResponse({"status": "User is already an agent"})
    if float(user.wallet) < float(agent_price):
        return JsonResponse({"status": "Insufficient Balance", "icon": "error"})
    user.wallet -= float(agent_price)
    user.status = "Agent"
    user.save()
    new_registration = models.AgentRegistration.objects.create(
        amount=agent_price,
        user=user
    )
    new_registration.save()
    return JsonResponse({"status": "Registration Successful"})


@login_required(login_url='login')
def services(request):
    return render(request, "layouts/services.html")


def t_and_c(request):
    return render(request, "layouts/about.html")


@ratelimit(key='ip', rate='10/m')
@login_required(login_url='login')
def pay_with_wallet(request):
    if request.method == "POST":
        admin = models.AdminInfo.objects.filter().first().phone_number
        user = models.CustomUser.objects.get(id=request.user.id)
        phone_number = request.POST.get("phone")
        amount = request.POST.get("amount")
        reference = request.POST.get("reference")
        if user.wallet is None:
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge. Admin Contact Info: 0{admin}'})
        elif user.wallet <= 0 or user.wallet < float(amount):
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge. Admin Contact Info: 0{admin}'})
        print(phone_number)
        print(amount)
        print(reference)

        if user.status == "User":
            bundle = models.IshareBundlePrice.objects.get(price=float(amount)).bundle_volume
        elif user.status == "Agent":
            bundle = models.AgentIshareBundlePrice.objects.get(price=float(amount)).bundle_volume
        elif user.status == "Super Agent":
            bundle = models.SuperAgentIshareBundlePrice.objects.get(price=float(amount)).bundle_volume
        else:
            bundle = models.IshareBundlePrice.objects.get(price=float(amount)).bundle_volume

        print(bundle)
        send_bundle_response = helper.send_bundle(request.user, phone_number, bundle, reference)
        print(send_bundle_response)

        sms_headers = {
            'Authorization': 'Bearer 1136|LwSl79qyzTZ9kbcf9SpGGl1ThsY0Ujf7tcMxvPze',
            'Content-Type': 'application/json'
        }

        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
        if send_bundle_response != "bad response":
            print("good response")
            if send_bundle_response["data"]["request_status_code"] == "200" or send_bundle_response[
                "request_message"] == "Successful":
                new_transaction = models.IShareBundleTransaction.objects.create(
                    user=request.user,
                    bundle_number=phone_number,
                    offer=f"{bundle}MB",
                    reference=reference,
                    transaction_status="Completed"
                )
                new_transaction.save()
                user.wallet -= float(amount)
                user.wallet = float(user.wallet)
                user.save()

                new_profit_instance = models.ProfitInstance.objects.create(
                    selling_price_total=amount,
                    channel="AT",  # Set your channel here based on user status
                )
                new_profit_instance.save()

                new_wallet_transaction = models.WalletTransaction.objects.create(
                    user=request.user,
                    transaction_type="Debit",
                    transaction_amount=float(amount),
                    transaction_use="AT",
                    new_balance=user.wallet
                )
                new_wallet_transaction.save()
                receiver_message = f"Your bundle purchase has been completed successfully. {bundle}MB has been credited to you by {request.user.phone}.\nReference: {reference}\n"
                sms_message = f"Hello @{request.user.username}. Your bundle purchase has been completed successfully. {bundle}MB has been credited to {phone_number}.\nReference: {reference}\nCurrent Wallet Balance: {user.wallet}\nThank you for using Geosams.\n\nGeosams"

                # num_without_0 = phone_number[1:]
                # print(num_without_0)
                # receiver_body = {
                #     'recipient': f"233{num_without_0}",
                #     'sender_id': 'Geosams',
                #     'message': receiver_message
                # }
                #
                # response = requests.request('POST', url=sms_url, params=receiver_body, headers=sms_headers)
                # print(response.text)
                #
                # sms_body = {
                #     'recipient': f"233{request.user.phone}",
                #     'sender_id': 'Geosams',
                #     'message': sms_message
                # }
                #
                # response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
                #
                # print(response.text)

                response1 = requests.get(
                    f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=UnBzemdvanJyUGxhTlJzaVVQaHk&to=0{request.user.phone}&from=GEO_AT&sms={sms_message}")
                print(response1.text)

                response2 = requests.get(
                    f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=UnBzemdvanJyUGxhTlJzaVVQaHk&to={phone_number}&from=GEO_AT&sms={receiver_message}")
                print(response2.text)
                return JsonResponse({'status': 'Transaction Completed Successfully', 'icon': 'success'})
            else:
                new_transaction = models.IShareBundleTransaction.objects.create(
                    user=request.user,
                    bundle_number=phone_number,
                    offer=f"{bundle}MB",
                    reference=reference,
                    transaction_status="Failed"
                )
                new_transaction.save()
                return JsonResponse({'status': 'Something went wrong', 'icon': 'error'})
    return redirect('airtel-tigo')


@ratelimit(key='ip', rate='10/m')
@login_required(login_url='login')
def airtel_tigo(request):
    user = models.CustomUser.objects.get(id=request.user.id)
    status = user.status
    form = forms.IShareBundleForm(status)
    reference = helper.ref_generator()
    user_email = request.user.email
    if request.method == "POST":
        form = forms.IShareBundleForm(data=request.POST, status=status)
        if form.is_valid():
            phone_number = form.cleaned_data["phone_number"]
            amount = form.cleaned_data["offers"]

            print(amount.price)

            details = {
                'phone_number': phone_number,
                'offers': amount.price
            }

            new_payment = models.Payment.objects.create(
                user=request.user,
                reference=reference,
                transaction_date=datetime.now(),
                transaction_details=details,
                channel="ishare",
            )
            new_payment.save()
            print("payment saved")
            print("form valid")

            url = "https://payproxyapi.hubtel.com/items/initiate"

            payload = json.dumps({
                "totalAmount": float(amount.price) + (1 / 100) * float(amount.price),
                "description": "N/A",
                "callbackUrl": "https://www.geosams.com/hubtel_webhook",
                "returnUrl": "https://www.geosams.com",
                "cancellationUrl": "https://www.geosams.com",
                "merchantAccountNumber": "2021482",
                "clientReference": new_payment.reference
            })
            headers = {
                'Content-Type': 'application/json',
                'Authorization': config("HUBTEL_TOKEN")
            }

            response = requests.request("POST", url, headers=headers, data=payload)

            data = response.json()

            checkoutUrl = data['data']['checkoutUrl']

            return redirect(checkoutUrl)
    # if request.method == "POST":
    #     form = forms.IShareBundleForm(data=request.POST, status=status)
    #     payment_reference = request.POST.get("reference")
    #     amount_paid = request.POST.get("amount")
    #     new_payment = models.Payment.objects.create(
    #         user=request.user,
    #         reference=payment_reference,
    #         amount=amount_paid,
    #         transaction_date=datetime.now(),
    #         transaction_status="Completed"
    #     )
    #     new_payment.save()
    #     print("payment saved")
    #     print("form valid")
    #     phone_number = request.POST.get("phone")
    #     offer = request.POST.get("amount")
    #     print(offer)
    #     if user.status == "User":
    #         bundle = models.IshareBundlePrice.objects.get(price=float(offer)).bundle_volume
    #     elif user.status == "Agent":
    #         bundle = models.AgentIshareBundlePrice.objects.get(price=float(offer)).bundle_volume
    #     elif user.status == "Super Agent":
    #         bundle = models.SuperAgentIshareBundlePrice.objects.get(price=float(offer)).bundle_volume
    #     else:
    #         bundle = models.IshareBundlePrice.objects.get(price=float(offer)).bundle_volume
    #
    #     new_transaction = models.IShareBundleTransaction.objects.create(
    #         user=request.user,
    #         bundle_number=phone_number,
    #         offer=f"{bundle}MB",
    #         reference=payment_reference,
    #         transaction_status="Pending"
    #     )
    #     print("created")
    #     new_transaction.save()
    #
    #     print("===========================")
    #     print(phone_number)
    #     print(bundle)
    #     send_bundle_response = helper.send_bundle(request.user, phone_number, bundle, payment_reference)
    #     data = send_bundle_response.json()
    #
    #     print(data)
    #
    #     sms_headers = {
    #         'Authorization': 'Bearer 1136|LwSl79qyzTZ9kbcf9SpGGl1ThsY0Ujf7tcMxvPze',
    #         'Content-Type': 'application/json'
    #     }
    #
    #     sms_url = 'https://webapp.usmsgh.com/api/sms/send'
    #
    #     if send_bundle_response.status_code == 200:
    #         if data["code"] == "0000":
    #             transaction_to_be_updated = models.IShareBundleTransaction.objects.get(reference=payment_reference)
    #             print("got here")
    #             print(transaction_to_be_updated.transaction_status)
    #             transaction_to_be_updated.transaction_status = "Completed"
    #             transaction_to_be_updated.save()
    #             print(request.user.phone)
    #             print("***********")
    #             receiver_message = f"Your bundle purchase has been completed successfully. {bundle}MB has been credited to you by {request.user.phone}.\nReference: {payment_reference}\n"
    #             sms_message = f"Hello @{request.user.username}. Your bundle purchase has been completed successfully. {bundle}MB has been credited to {phone_number}.\nReference: {payment_reference}\nThank you for using Geosams GH.\n\nThe Geosams GH"
    #
    #             num_without_0 = phone_number[1:]
    #             print(num_without_0)
    #             receiver_body = {
    #                 'recipient': f"233{num_without_0}",
    #                 'sender_id': 'Geosams',
    #                 'message': receiver_message
    #             }
    #
    #             response = requests.request('POST', url=sms_url, params=receiver_body, headers=sms_headers)
    #             print(response.text)
    #
    #             sms_body = {
    #                 'recipient': f"233{request.user.phone}",
    #                 'sender_id': 'Geosams',
    #                 'message': sms_message
    #             }
    #
    #             response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
    #
    #             print(response.text)
    #
    #             return JsonResponse({'status': 'Transaction Completed Successfully', 'icon': 'success'})
    #         else:
    #             transaction_to_be_updated = models.IShareBundleTransaction.objects.get(reference=payment_reference)
    #             transaction_to_be_updated.transaction_status = "Failed"
    #             new_transaction.save()
    #             sms_message = f"Hello @{request.user.username}. Something went wrong with your transaction. Contact us for enquiries.\nBundle: {bundle}MB\nPhone Number: {phone_number}.\nReference: {payment_reference}\nThank you for using Geosams GH.\n\nThe Geosams GH"
    #
    #             sms_body = {
    #                 'recipient': f"233{request.user.phone}",
    #                 'sender_id': 'Geosams',
    #                 'message': sms_message
    #             }
    #             # response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
    #             # print(response.text)
    #             # r_sms_url = f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=UmpEc1JzeFV4cERKTWxUWktqZEs&to={phone_number}&from=Geosams GH&sms={receiver_message}"
    #             # response = requests.request("GET", url=r_sms_url)
    #             # print(response.text)
    #             return JsonResponse({'status': 'Something went wrong', 'icon': 'error'})
    #     else:
    #         transaction_to_be_updated = models.IShareBundleTransaction.objects.get(reference=payment_reference)
    #         transaction_to_be_updated.transaction_status = "Failed"
    #         new_transaction.save()
    #         sms_message = f"Hello @{request.user.username}. Something went wrong with your transaction. Contact us for enquiries.\nBundle: {bundle}MB\nPhone Number: {phone_number}.\nReference: {payment_reference}\nThank you for using Geosams GH.\n\nThe Geosams GH"
    #
    #         sms_body = {
    #             'recipient': f'233{request.user.phone}',
    #             'sender_id': 'Geosams',
    #             'message': sms_message
    #         }
    #
    #         # response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
    #         #
    #         # print(response.text)
    #         return JsonResponse({'status': 'Something went wrong', 'icon': 'error'})
    user = models.CustomUser.objects.get(id=request.user.id)
    context = {"form": form, "ref": reference, "email": user_email,
               "wallet": 0 if user.wallet is None else user.wallet}
    return render(request, "layouts/services/at.html", context=context)


@ratelimit(key='ip', rate='10/m')
@login_required(login_url='login')
def mtn_pay_with_wallet(request):
    if request.method == "POST":
        admin = models.AdminInfo.objects.filter().first().phone_number
        user = models.CustomUser.objects.get(id=request.user.id)
        phone_number = request.POST.get("phone")
        amount = request.POST.get("amount")
        reference = request.POST.get("reference")
        print(phone_number)
        print(amount)
        print(reference)
        sms_headers = {
            'Authorization': 'Bearer 1136|LwSl79qyzTZ9kbcf9SpGGl1ThsY0Ujf7tcMxvPze',
            'Content-Type': 'application/json'
        }

        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
        admin = models.AdminInfo.objects.filter().first().phone_number

        if user.wallet is None:
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge. Admin Contact Info: 0{admin}'})
        elif user.wallet <= 0 or user.wallet < float(amount):
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge. Admin Contact Info: 0{admin}'})
        if user.status == "User":
            bundle = models.MTNBundlePrice.objects.get(price=float(amount)).bundle_volume
        elif user.status == "Agent":
            bundle = models.AgentMTNBundlePrice.objects.get(price=float(amount)).bundle_volume
        elif user.status == "Super Agent":
            bundle = models.SuperAgentMTNBundlePrice.objects.get(price=float(amount)).bundle_volume
        else:
            bundle = models.MTNBundlePrice.objects.get(price=float(amount)).bundle_volume
        print(bundle)
        sms_message = f"An order has been placed. {bundle}MB for {phone_number}"
        new_mtn_transaction = models.MTNTransaction.objects.create(
            user=request.user,
            bundle_number=phone_number,
            offer=f"{bundle}MB",
            reference=reference,
            amount=amount
        )
        new_mtn_transaction.save()
        user.wallet -= float(amount)
        user.wallet = float(user.wallet)
        user.save()

        new_profit_instance = models.ProfitInstance.objects.create(
            selling_price_total=amount,
            channel="MTN",  # Set your channel here based on user status
        )
        new_profit_instance.save()

        new_wallet_transaction = models.WalletTransaction.objects.create(
            user=request.user,
            transaction_type="Debit",
            transaction_amount=float(amount),
            transaction_use="MTN",
            new_balance=user.wallet
        )
        new_wallet_transaction.save()
        sms_body = {
            'recipient': f"233{admin}",
            'sender_id': 'Geosams',
            'message': sms_message
        }
        response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
        print(response.text)
        return JsonResponse({'status': "Your transaction will be completed shortly", 'icon': 'success'})
    return redirect('mtn')


@login_required(login_url='login')
def voda_pay_with_wallet(request):
    if request.method == "POST":
        user = models.CustomUser.objects.get(id=request.user.id)
        phone_number = request.POST.get("phone")
        amount = request.POST.get("amount")
        reference = request.POST.get("reference")
        print(phone_number)
        print(amount)
        print(reference)
        if user.wallet is None:
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge.'})
        elif user.wallet <= 0 or user.wallet < float(amount):
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge.'})
        if user.status == "User":
            bundle = models.VodaBundlePrice.objects.get(price=float(amount)).bundle_volume
        elif user.status == "Agent":
            bundle = models.AgentVodaBundlePrice.objects.get(price=float(amount)).bundle_volume
        elif user.status == "Super Agent":
            bundle = models.SuperAgentVodaBundlePrice.objects.get(price=float(amount)).bundle_volume
        else:
            bundle = models.VodaBundlePrice.objects.get(price=float(amount)).bundle_volume

        print(bundle)
        new_mtn_transaction = models.VodafoneTransaction.objects.create(
            user=request.user,
            bundle_number=phone_number,
            offer=f"{bundle}MB",
            reference=reference,
        )
        new_mtn_transaction.save()

        user.wallet -= float(amount)
        user.save()

        new_wallet_transaction = models.WalletTransaction.objects.create(
            user=request.user,
            transaction_type="Debit",
            transaction_amount=float(amount),
            transaction_use="Telecel",
            new_balance=user.wallet
        )
        new_wallet_transaction.save()

        new_profit_instance = models.ProfitInstance.objects.create(
            selling_price_total=amount,
            channel="Telecel",  # Set your channel here based on user status
        )
        new_profit_instance.save()
        return JsonResponse({'status': "Your transaction will be completed shortly", 'icon': 'success'})
    return redirect('voda')


@ratelimit(key='ip', rate='10/m')
@login_required(login_url='login')
def big_time_pay_with_wallet(request):
    if request.method == "POST":
        user = models.CustomUser.objects.get(id=request.user.id)
        phone_number = request.POST.get("phone")
        amount = request.POST.get("amount")
        reference = request.POST.get("reference")
        print(phone_number)
        print(amount)
        print(reference)
        if user.wallet is None:
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge.'})
        elif user.wallet <= 0 or user.wallet < float(amount):
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge.'})
        if user.status == "User":
            bundle = models.BigTimeBundlePrice.objects.get(price=float(amount)).bundle_volume
        elif user.status == "Agent":
            bundle = models.AgentBigTimeBundlePrice.objects.get(price=float(amount)).bundle_volume
        elif user.status == "Super Agent":
            bundle = models.SuperAgentBigTimeBundlePrice.objects.get(price=float(amount)).bundle_volume
        else:
            bundle = models.BigTimeBundlePrice.objects.get(price=float(amount)).bundle_volume
        print(bundle)
        new_mtn_transaction = models.BigTimeTransaction.objects.create(
            user=request.user,
            bundle_number=phone_number,
            offer=f"{bundle}MB",
            reference=reference,
        )
        new_mtn_transaction.save()
        user.wallet -= float(amount)
        user.wallet = float(user.wallet)
        user.save()
        new_wallet_transaction = models.WalletTransaction.objects.create(
            user=request.user,
            transaction_type="Debit",
            transaction_amount=float(amount),
            transaction_use="AT BigTime",
            new_balance=user.wallet
        )
        new_wallet_transaction.save()

        new_profit_instance = models.ProfitInstance.objects.create(
            selling_price_total=amount,
            channel="BigTime",  # Set your channel here based on user status
        )
        new_profit_instance.save()

        return JsonResponse({'status': "Your transaction will be completed shortly", 'icon': 'success'})
    return redirect('big_time')


@ratelimit(key='ip', rate='10/m')
@login_required(login_url='login')
def mtn(request):
    user = models.CustomUser.objects.get(id=request.user.id)
    status = user.status
    form = forms.MTNForm(status=status)
    reference = helper.ref_generator()
    user_email = request.user.email
    admin = models.AdminInfo.objects.filter().first().phone_number
    if request.method == "POST":
        form = forms.MTNForm(data=request.POST, status=status)
        if form.is_valid():
            phone_number = form.cleaned_data['phone_number']
            amount = form.cleaned_data['offers']

            print(amount.price)

            details = {
                'phone_number': phone_number,
                'offers': amount.price
            }

            new_payment = models.Payment.objects.create(
                user=request.user,
                reference=reference,
                transaction_date=datetime.now(),
                transaction_details=details,
                channel="mtn",
            )
            new_payment.save()
            print("payment saved")
            print("form valid")

            url = "https://payproxyapi.hubtel.com/items/initiate"

            payload = json.dumps({
                "totalAmount": float(amount.price) + (1 / 100) * float(amount.price),
                "description": "N/A",
                "callbackUrl": "https://www.geosams.com/hubtel_webhook",
                "returnUrl": "https://www.geosams.com",
                "cancellationUrl": "https://www.geosams.com",
                "merchantAccountNumber": "2021482",
                "clientReference": new_payment.reference
            })
            headers = {
                'Content-Type': 'application/json',
                'Authorization': config("HUBTEL_TOKEN")
            }

            response = requests.request("POST", url, headers=headers, data=payload)

            data = response.json()

            checkoutUrl = data['data']['checkoutUrl']

            return redirect(checkoutUrl)
    user = models.CustomUser.objects.get(id=request.user.id)
    try:
        api_user = models.MTNAPIUsers.objects.filter(user=user).first()
        api_wallet = api_user.wallet_balance
        context = {'form': form, "ref": reference, "email": user_email,
                   "wallet": 0 if user.wallet is None else user.wallet, 'api_wallet': api_wallet}
        return render(request, "layouts/services/mtn.html", context=context)
    except:
        context = {'form': form, "ref": reference, "email": user_email,
                   "wallet": 0 if user.wallet is None else user.wallet}
        return render(request, "layouts/services/mtn.html", context=context)


@login_required(login_url='login')
def voda(request):
    user = models.CustomUser.objects.get(id=request.user.id)
    status = user.status
    form = forms.VodaBundleForm(status)
    reference = helper.ref_generator()
    user_email = request.user.email

    if request.method == "POST":
        payment_reference = request.POST.get("reference")
        amount_paid = request.POST.get("amount")
        new_payment = models.Payment.objects.create(
            user=request.user,
            reference=payment_reference,
            amount=amount_paid,
            transaction_date=datetime.now(),
            transaction_status="Pending"
        )
        new_payment.save()
        phone_number = request.POST.get("phone")
        offer = request.POST.get("amount")
        if user.status == "User":
            bundle = models.VodaBundlePrice.objects.get(price=float(offer)).bundle_volume
        elif user.status == "Agent":
            bundle = models.AgentVodaBundlePrice.objects.get(price=float(offer)).bundle_volume
        elif user.status == "Super Agent":
            bundle = models.SuperAgentVodaBundlePrice.objects.get(price=float(offer)).bundle_volume
        else:
            bundle = models.VodaBundlePrice.objects.get(price=float(offer)).bundle_volume

        print(phone_number)
        new_mtn_transaction = models.VodafoneTransaction.objects.create(
            user=request.user,
            bundle_number=phone_number,
            offer=f"{bundle}MB",
            reference=payment_reference,
        )
        new_mtn_transaction.save()
        return JsonResponse({'status': "Your transaction will be completed shortly", 'icon': 'success'})
    user = models.CustomUser.objects.get(id=request.user.id)
    # phone_num = user.phone
    # mtn_dict = {}
    #
    # if user.status == "Agent":
    #     mtn_offer = models.AgentMTNBundlePrice.objects.all()
    # else:
    #     mtn_offer = models.MTNBundlePrice.objects.all()
    # for offer in mtn_offer:
    #     mtn_dict[str(offer)] = offer.bundle_volume
    context = {'form': form,
               "ref": reference, "email": user_email, "wallet": 0 if user.wallet is None else user.wallet}
    return render(request, "layouts/services/voda.html", context=context)


@ratelimit(key='ip', rate='10/m')
@login_required(login_url='login')
def afa_registration(request):
    user = models.CustomUser.objects.get(id=request.user.id)
    reference = helper.ref_generator()
    price = models.AdminInfo.objects.filter().first().afa_price
    if user.status == "Super Agent":
        price = models.AdminInfo.objects.filter().first().afa_super_agent_price
    user_email = request.user.email
    print(price)
    if request.method == "POST":
        form = forms.AFARegistrationForm(request.POST)
        if form.is_valid():
            # name = transaction_details["name"]
            # phone_number = transaction_details["phone"]
            # gh_card_number = transaction_details["card"]
            # occupation = transaction_details["occupation"]
            # date_of_birth = transaction_details["date_of_birth"]
            details = {
                "name": form.cleaned_data["name"],
                "phone": form.cleaned_data["phone_number"],
                "card": form.cleaned_data["gh_card_number"],
                "occupation": form.cleaned_data["occupation"],
                "date_of_birth": form.cleaned_data["date_of_birth"],
            }
            new_payment = models.Payment.objects.create(
                user=request.user,
                reference=reference,
                transaction_details=details,
                transaction_date=datetime.now(),
                channel="afa"
            )
            new_payment.save()

            url = "https://payproxyapi.hubtel.com/items/initiate"

            payload = json.dumps({
                "totalAmount": float(price) + (1 / 100) * float(price),
                "description": "N/A",
                "callbackUrl": "https://www.geosams.com/hubtel_webhook",
                "returnUrl": "https://www.geosams.com",
                "cancellationUrl": "https://www.geosams.com",
                "merchantAccountNumber": "2021482",
                "clientReference": new_payment.reference
            })
            headers = {
                'Content-Type': 'application/json',
                'Authorization': config("HUBTEL_TOKEN")
            }

            response = requests.request("POST", url, headers=headers, data=payload)

            data = response.json()

            checkoutUrl = data['data']['checkoutUrl']

            return redirect(checkoutUrl)
    form = forms.AFARegistrationForm()
    context = {'form': form, 'ref': reference, 'price': price, "email": user_email,
               "wallet": 0 if user.wallet is None else user.wallet}
    return render(request, "layouts/services/afa.html", context=context)


@ratelimit(key='ip', rate='10/m')
@login_required(login_url='login')
def afa_registration_wallet(request):
    if request.method == "POST":
        user = models.CustomUser.objects.get(id=request.user.id)
        phone_number = request.POST.get("phone")
        amount = request.POST.get("amount")
        reference = request.POST.get("reference")
        name = request.POST.get("name")
        card_number = request.POST.get("card")
        occupation = request.POST.get("occupation")
        date_of_birth = request.POST.get("birth")
        price = models.AdminInfo.objects.filter().first().afa_price
        if user.status == "Super Agent":
            price = models.AdminInfo.objects.filter().first().afa_super_agent_price

        if user.wallet is None:
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge.'})
        elif user.wallet <= 0 or user.wallet < float(amount):
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge.'})

        new_registration = models.AFARegistration.objects.create(
            user=user,
            reference=reference,
            name=name,
            phone_number=phone_number,
            gh_card_number=card_number,
            occupation=occupation,
            date_of_birth=date_of_birth
        )
        new_registration.save()
        user.wallet -= float(price)
        user.wallet = float(user.wallet)
        user.save()
        new_wallet_transaction = models.WalletTransaction.objects.create(
            user=request.user,
            transaction_type="Debit",
            transaction_amount=float(amount),
            transaction_use="Afa",
            new_balance=user.wallet
        )
        new_wallet_transaction.save()

        # new_profit_instance = models.ProfitInstance.objects.create(
        #     selling_price_total=amount,
        #     channel="MTN",  # Set your channel here based on user status
        # )
        # new_profit_instance.save()
        return JsonResponse({'status': "Your transaction will be completed shortly", 'icon': 'success'})
    return redirect('home')


@login_required(login_url='login')
def big_time(request):
    user = models.CustomUser.objects.get(id=request.user.id)
    status = user.status
    form = forms.BigTimeBundleForm(status)
    reference = helper.ref_generator()
    user_email = request.user.email

    if request.method == "POST":
        form = forms.BigTimeBundleForm(data=request.POST, status=status)
        if form.is_valid():
            phone_number = form.cleaned_data['phone_number']
            amount = form.cleaned_data['offers']
            details = {
                'phone_number': phone_number,
                'offers': amount.price
            }
            new_payment = models.Payment.objects.create(
                user=request.user,
                reference=reference,
                transaction_details=details,
                transaction_date=datetime.now(),
                channel="bigtime"
            )
            new_payment.save()

            url = "https://payproxyapi.hubtel.com/items/initiate"

            payload = json.dumps({
                "totalAmount": float(amount.price) + (1 / 100) * float(amount.price),
                "description": "N/A",
                "callbackUrl": "https://www.geosams.com/hubtel_webhook",
                "returnUrl": "https://www.geosams.com",
                "cancellationUrl": "https://www.geosams.com",
                "merchantAccountNumber": "2021482",
                "clientReference": new_payment.reference
            })
            headers = {
                'Content-Type': 'application/json',
                'Authorization': config("HUBTEL_TOKEN")
            }

            response = requests.request("POST", url, headers=headers, data=payload)

            data = response.json()

            checkoutUrl = data['data']['checkoutUrl']

            return redirect(checkoutUrl)
    user = models.CustomUser.objects.get(id=request.user.id)
    # phone_num = user.phone
    # mtn_dict = {}
    #
    # if user.status == "Agent":
    #     mtn_offer = models.AgentMTNBundlePrice.objects.all()
    # else:
    #     mtn_offer = models.MTNBundlePrice.objects.all()
    # for offer in mtn_offer:
    #     mtn_dict[str(offer)] = offer.bundle_volume
    context = {'form': form,
               "ref": reference, "email": user_email, "wallet": 0 if user.wallet is None else user.wallet}
    return render(request, "layouts/services/big_time.html", context=context)


@login_required(login_url='login')
def history(request):
    user_transactions = models.IShareBundleTransaction.objects.filter(user=request.user).order_by(
        'transaction_date').reverse()[:200]
    header = "AirtelTigo Transactions"
    net = "tigo"
    try:
        user = models.CustomUser.objects.get(id=request.user.id)
        api_user = models.MTNAPIUsers.objects.filter(user=user).first()
        context = {'txns': user_transactions, "header": header, "net": net, "api_user": api_user}
        return render(request, "layouts/history.html", context=context)
    except:
        context = {'txns': user_transactions, "header": header, "net": net}
        return render(request, "layouts/history.html", context=context)


@login_required(login_url='login')
def wallet_history(request):
    user_wallet_transactions = models.WalletTransaction.objects.filter(user=request.user).order_by(
        'transaction_date').reverse()[:500]
    header = "Wallet Transactions"
    net = "wallet"
    context = {'txns': user_wallet_transactions, "header": header, "net": net}
    return render(request, "layouts/wallet_history.html", context=context)


@login_required(login_url='login')
def api_wallet_history(request):
    user_wallet_transactions = models.ApiWalletTransaction.objects.filter(user=request.user,
                                                                          transaction_channel="MTN").order_by(
        'transaction_date').reverse()[:500]
    header = "API Wallet Transactions"
    net = "api_wallet"
    context = {'txns': user_wallet_transactions, "header": header, "net": net}
    return render(request, "layouts/api_wallet_history.html", context=context)


@login_required(login_url='login')
def telecel_api_wallet_history(request):
    user_wallet_transactions = models.ApiWalletTransaction.objects.filter(user=request.user,
                                                                          transaction_channel="Telecel").order_by(
        'transaction_date').reverse()[:500]
    header = "Telecel API Wallet Transactions"
    net = "telecel_api_wallet"
    context = {'txns': user_wallet_transactions, "header": header, "net": net}
    return render(request, "layouts/voda_api_wallet_history.html", context=context)


@login_required(login_url='login')
def mtn_history(request):
    user_transactions = models.MTNTransaction.objects.filter(user=request.user).order_by(
        'transaction_date').reverse()[:800]
    header = "MTN Transactions"
    net = "mtn"
    context = {'txns': user_transactions, "header": header, "net": net}
    return render(request, "layouts/history.html", context=context)


@login_required(login_url='login')
def voda_history(request):
    user_transactions = models.VodafoneTransaction.objects.filter(user=request.user).order_by(
        'transaction_date').reverse()[:200]
    header = "Vodafone Transactions"
    net = "voda"
    context = {'txns': user_transactions, "header": header, "net": net}
    return render(request, "layouts/history.html", context=context)


from django.contrib import messages
from django.shortcuts import redirect, render
from django.http import HttpResponse
import csv
from io import StringIO
import datetime
from . import models


@login_required(login_url='login')
def admin_voda_history(request, status):
    if request.user.is_staff and request.user.is_superuser:
        if request.method == "POST":
            uploaded_file = request.FILES.get('file')
            if not uploaded_file:
                messages.error(request, "No CSV file found")
                return redirect('voda_admin', status=status)

            # Read the uploaded CSV file into memory
            csv_buffer = StringIO(uploaded_file.read().decode('utf-8'))
            reader = csv.reader(csv_buffer)

            # Skip the header if present
            header = next(reader)

            # Load the existing data from the CSV file into a list
            csv_data = list(reader)

            # Query your Django model
            queryset = models.VodafoneTransaction.objects.filter(transaction_status="Pending")

            # Assuming we have identified the recipient and data column indices
            recipient_col_index = 0  # Adjust based on actual index in the CSV
            data_col_index = 1  # Adjust based on actual index in the CSV

            # Update the CSV data with the queryset data
            for record in queryset:
                recipient_value = f"0{record.bundle_number}"  # Ensure it's a string to preserve formatting
                data_value = record.offer  # Adjust based on actual field type
                cleaned_data_value = float(data_value.replace('MB', ''))
                data_value_gb = round(float(cleaned_data_value) / 1000, 2)

                # Append the new data to the CSV data list
                csv_data.append([recipient_value, data_value_gb])

                # Update the record status, if necessary
                record.transaction_status = "Processing"
                record.save()

            # Create a new CSV buffer to write the updated data
            csv_output = StringIO()
            writer = csv.writer(csv_output)

            # Write the header if present
            writer.writerow(header)

            # Write the updated data to the CSV buffer
            writer.writerows(csv_data)

            # Prepare the response with the modified CSV file
            response = HttpResponse(csv_output.getvalue(), content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename={}.csv'.format(
                datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S"))

            return response

        all_txns = models.VodafoneTransaction.objects.filter(transaction_status=status).order_by('-transaction_date')[
                   :800]
        context = {'txns': all_txns, 'status': status}
        return render(request, "layouts/services/voda_admin.html", context=context)
    else:
        messages.error(request, "Access Denied")
        return redirect('voda_admin', status=status)


@login_required(login_url='login')
def voda_change_excel_status(request, status, to_change_to):
    transactions = models.VodafoneTransaction.objects.filter(
        transaction_status=status) if to_change_to != "Completed" else models.VodafoneTransaction.objects.filter(
        transaction_status=status).order_by('transaction_date')
    for transaction in transactions:
        transaction.transaction_status = to_change_to
        transaction.save()
    messages.success(request, f"Status changed from {status} to {to_change_to}")
    return redirect("voda_admin", status=status)


@login_required(login_url='login')
def big_time_history(request):
    user_transactions = models.BigTimeTransaction.objects.filter(user=request.user).order_by(
        'transaction_date').reverse()[:200]
    header = "Big Time Transactions"
    net = "bt"
    context = {'txns': user_transactions, "header": header, "net": net}
    return render(request, "layouts/history.html", context=context)


@login_required(login_url='login')
def afa_history(request):
    user_transactions = models.AFARegistration.objects.filter(user=request.user).order_by(
        'transaction_date').reverse()[:200]
    header = "AFA Registrations"
    net = "afa"
    context = {'txns': user_transactions, "header": header, "net": net}
    return render(request, "layouts/afa_history.html", context=context)


def verify_transaction(request, reference):
    if request.method == "GET":
        response = helper.verify_paystack_transaction(reference)
        data = response.json()
        try:
            status = data["data"]["status"]
            amount = data["data"]["amount"]
            api_reference = data["data"]["reference"]
            date = data["data"]["paid_at"]
            real_amount = float(amount) / 100
            print(status)
            print(real_amount)
            print(api_reference)
            print(reference)
            print(date)
        except:
            status = data["status"]
        return JsonResponse({'status': status})


# @login_required(login_url='login')
# def admin_mtn_history(request):
#     if request.user.is_staff and request.user.is_superuser:
#         all_txns = models.MTNTransaction.objects.all().order_by('transaction_date').reverse()[:1000]
#         context = {'txns': all_txns}
#         return render(request, "layouts/services/mtn_admin.html", context=context)


@login_required(login_url='login')
def change_excel_status(request, status, to_change_to):
    transactions = models.MTNTransaction.objects.filter(
        transaction_status=status) if to_change_to != "Completed" else models.MTNTransaction.objects.filter(
        transaction_status=status).order_by('transaction_date')
    for transaction in transactions:
        transaction.transaction_status = to_change_to
        transaction.save()
        if to_change_to == "Completed":
            # transaction_number = transaction.user.phone
            # sms_headers = {
            #     'Authorization': 'Bearer 1050|VDqcCUHwCBEbjcMk32cbdOhCFlavpDhy6vfgM4jU',
            #     'Content-Type': 'application/json'
            # }
            #
            # sms_url = 'https://webapp.usmsgh.com/api/sms/send'
            # sms_message = f"Your MTN transaction has been completed. {transaction.bundle_number} has been credited with {transaction.offer}.\nTransaction Reference: {transaction.reference}"
            #
            # sms_body = {
            #     'recipient': f"233{transaction_number}",
            #     'sender_id': 'Geosams',
            #     'message': sms_message
            # }
            # # try:
            # #     response1 = requests.get(
            # #         f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=UnBzemdvanJyUGxhTlJzaVVQaHk&to=0{transaction_number}&from=GEO_AT&sms={sms_message}")
            # #     print(response1.text)
            # # except:
            messages.success(request, f"Transactions Completed")
            return redirect('mtn_admin', status=status)
        else:
            messages.success(request, f"Status changed from {status} to {to_change_to}")
            return redirect("mtn_admin", status=status)
    messages.success(request, f"Status changed from {status} to {to_change_to}")
    return redirect("mtn_admin", status=status)


@login_required(login_url='login')
def admin_mtn_history(request, status):
    if request.user.is_staff and request.user.is_superuser:
        if request.method == "POST":
            from io import BytesIO
            from openpyxl import load_workbook
            from django.http import HttpResponse
            import datetime

            # Assuming `uploaded_file` is the Excel file uploaded by the user
            uploaded_file = request.FILES['file'] if 'file' in request.FILES else None
            if not uploaded_file:
                messages.error(request, "No excel file found")
                return redirect('mtn_admin', status=status)

            # Load the uploaded Excel file into memory
            excel_buffer = BytesIO(uploaded_file.read())
            book = load_workbook(excel_buffer)
            sheet = book.active  # Assuming the data is on the active sheet

            # Assuming we have identified the recipient and data column indices
            # Replace these with the actual indices if available
            recipient_col_index = 1  # Example index for "RECIPIENT"
            data_col_index = 2  # Example index for "DATA"

            # Query your Django model
            queryset = models.MTNTransaction.objects.filter(transaction_status="Pending")

            # Determine the starting row for updates, preserving headers and any other pre-existing content
            start_row = 2  # Assuming data starts from row 2

            for record in queryset:
                # Assuming 'bundle_number' and 'offer' fields exist in your model
                recipient_value = f"0{record.bundle_number}"  # Ensure it's a string to preserve formatting
                data_value = record.offer  # Adjust based on actual field type
                cleaned_data_value = float(data_value.replace('MB', ''))
                data_value_gb = round(float(cleaned_data_value) / 1000, 2)

                # Find next available row (avoid overwriting non-empty rows if necessary)
                while sheet.cell(row=start_row, column=recipient_col_index).value is not None:
                    start_row += 1

                # Update cells
                sheet.cell(row=start_row, column=recipient_col_index, value=recipient_value)
                sheet.cell(row=start_row, column=data_col_index, value=data_value_gb)

                # Update the record status, if necessary
                record.transaction_status = "Processing"
                record.save()

            # Save the modified Excel file to the buffer
            excel_buffer.seek(0)  # Reset buffer position
            book.save(excel_buffer)

            # Prepare the response with the modified Excel file
            excel_buffer.seek(0)  # Reset buffer position to read the content
            response = HttpResponse(excel_buffer.getvalue(),
                                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename={}.xlsx'.format(
                datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S"))

            return response

        all_txns = models.MTNTransaction.objects.filter(transaction_status=status).order_by('-transaction_date')[:800]
        context = {'txns': all_txns, 'status': status}
        return render(request, "layouts/services/mtn_admin.html", context=context)
    else:
        messages.error(request, "Access Denied")
        return redirect('mtn_admin', status=status)


@login_required(login_url='login')
def admin_bt_history(request):
    if request.user.is_staff and request.user.is_superuser:
        all_txns = models.BigTimeTransaction.objects.filter().order_by('-transaction_date')[:1000]
        context = {'txns': all_txns}
        return render(request, "layouts/services/bt_admin.html", context=context)


@login_required(login_url='login')
def admin_afa_history(request):
    if request.user.is_staff and request.user.is_superuser:
        all_txns = models.AFARegistration.objects.filter().order_by('-transaction_date')[:1000]
        context = {'txns': all_txns}
        return render(request, "layouts/services/afa_admin.html", context=context)


@login_required(login_url='login')
def mark_as_sent(request, pk):
    if request.user.is_staff and request.user.is_superuser:
        txn = models.MTNTransaction.objects.filter(id=pk).first()
        print(txn)
        txn.transaction_status = "Completed"
        txn.save()
        sms_headers = {
            'Authorization': 'Bearer 1136|LwSl79qyzTZ9kbcf9SpGGl1ThsY0Ujf7tcMxvPze',
            'Content-Type': 'application/json'
        }

        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
        sms_message = f"Your account has been credited with {txn.offer}.\nTransaction Reference: {txn.reference}"

        sms_body = {
            'recipient': f"233{txn.bundle_number}",
            'sender_id': 'Geosams',
            'message': sms_message
        }
        response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
        print(response.text)
        return redirect('mtn_admin', status="Pending")


@login_required(login_url='login')
def bt_mark_as_sent(request, pk):
    if request.user.is_staff and request.user.is_superuser:
        txn = models.BigTimeTransaction.objects.filter(id=pk).first()
        print(txn)
        txn.transaction_status = "Completed"
        txn.save()
        sms_headers = {
            'Authorization': 'Bearer 1136|LwSl79qyzTZ9kbcf9SpGGl1ThsY0Ujf7tcMxvPze',
            'Content-Type': 'application/json'
        }

        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
        sms_message = f"Your AT BIG TIME transaction has been completed. {txn.bundle_number} has been credited with {txn.offer}.\nTransaction Reference: {txn.reference}"

        sms_body = {
            'recipient': f"233{txn.user.phone}",
            'sender_id': 'Geosams',
            'message': sms_message
        }
        try:
            response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
            print(response.text)
        except:
            messages.success(request, f"Transaction Completed")
            return redirect('bt_admin')
        messages.success(request, f"Transaction Completed")
        return redirect('bt_admin')


@login_required(login_url='login')
def voda_mark_as_sent(request, pk):
    if request.user.is_staff and request.user.is_superuser:
        txn = models.VodafoneTransaction.objects.filter(id=pk).first()
        print(txn)
        txn.transaction_status = "Completed"
        txn.save()
        sms_headers = {
            'Authorization': 'Bearer 1136|LwSl79qyzTZ9kbcf9SpGGl1ThsY0Ujf7tcMxvPze',
            'Content-Type': 'application/json'
        }

        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
        sms_message = f"Your Vodafone transaction has been completed. {txn.bundle_number} has been credited with {txn.offer}.\nTransaction Reference: {txn.reference}"

        sms_body = {
            'recipient': f"233{txn.user.phone}",
            'sender_id': 'Geosams',
            'message': sms_message
        }
        try:
            response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
            print(response.text)
        except Exception as e:
            print(e)
        messages.success(request, f"Transaction Completed")
        return redirect('voda_admin')


@login_required(login_url='login')
def afa_mark_as_sent(request, pk):
    if request.user.is_staff and request.user.is_superuser:
        txn = models.AFARegistration.objects.filter(id=pk).first()
        print(txn)
        txn.transaction_status = "Completed"
        txn.save()
        sms_headers = {
            'Authorization': 'Bearer 1136|LwSl79qyzTZ9kbcf9SpGGl1ThsY0Ujf7tcMxvPze',
            'Content-Type': 'application/json'
        }

        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
        sms_message = f"Your AFA Registration has been completed. {txn.phone_number} has been registered.\nTransaction Reference: {txn.reference}"

        sms_body = {
            'recipient': f"233{txn.user.phone}",
            'sender_id': 'Geosams',
            'message': sms_message
        }
        response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
        print(response.text)
        messages.success(request, f"Transaction Completed")
        return redirect('afa_admin')


@ratelimit(key='ip', rate='10/m')
@login_required(login_url='login')
def credit_user(request):
    form = forms.CreditUserForm()
    if request.user.is_superuser:
        if request.method == "POST":
            form = forms.CreditUserForm(request.POST)
            if form.is_valid():
                user = form.cleaned_data["user"]
                amount = form.cleaned_data["amount"]
                print(user)
                print(amount)
                user_needed = models.CustomUser.objects.get(username=user)
                if user_needed.wallet is None:
                    user_needed.wallet = float(amount)
                    user_needed.wallet = float(user.wallet)
                    new_wallet_transaction = models.WalletTransaction.objects.create(
                        user=user_needed,
                        transaction_type="Credit",
                        transaction_amount=float(amount),
                        transaction_use="Top up"
                    )
                else:
                    user_needed.wallet += float(amount)
                    user_needed.wallet = float(user.wallet)
                    new_wallet_transaction = models.WalletTransaction.objects.create(
                        user=request.user,
                        transaction_type="Credit",
                        transaction_amount=float(amount),
                        transaction_use="Top up"
                    )
                user_needed.save()
                new_wallet_transaction.new_balance = user_needed.wallet
                new_wallet_transaction.save()
                print(user_needed.username)
                messages.success(request, "Crediting Successful")
                return redirect('credit_user')
        context = {'form': form}
        return render(request, "layouts/services/credit.html", context=context)
    else:
        messages.error(request, "Access Denied")
        return redirect('home')


@login_required(login_url='login')
def topup_info(request):
    # if request.method == "POST":
    #     admin = models.AdminInfo.objects.filter().first().phone_number
    #     user = models.CustomUser.objects.get(id=request.user.id)
    #     amount = request.POST.get("amount")
    #
    #     if float(amount) < 50:
    #         messages.error(request, "Minimum amount is GHS 10")
    #         return redirect('topup-info')
    #     print(amount)
    #     reference = helper.top_up_ref_generator()
    #     details = {
    #         'topup_amount': amount
    #     }
    #     new_payment = models.Payment.objects.create(
    #         user=request.user,
    #         reference=reference,
    #         transaction_details=details,
    #         transaction_date=datetime.now(),
    #         channel="topup"
    #     )
    #     new_payment.save()
    #
    #     url = "https://payproxyapi.hubtel.com/items/initiate"
    #     print("hello world")
    #     print("Amount is " + amount)
    #
    #     try:
    #         total_amount = float(amount) + (1 / 100) * float(amount)
    #     except:
    #         return redirect('topup-info')
    #
    #     payload = json.dumps({
    #         "totalAmount": total_amount,
    #         "description": "N/A",
    #         "callbackUrl": "https://www.geosams.com/hubtel_webhook",
    #         "returnUrl": "https://www.geosams.com",
    #         "cancellationUrl": "https://www.geosams.com",
    #         "merchantAccountNumber": "2021482",
    #         "clientReference": new_payment.reference
    #     })
    #     headers = {
    #         'Content-Type': 'application/json',
    #         'Authorization': config("HUBTEL_TOKEN")
    #     }
    #
    #     response = requests.request("POST", url, headers=headers, data=payload)
    #
    #     data = response.json()
    #
    #     checkoutUrl = data['data']['checkoutUrl']
    #
    #     return redirect(checkoutUrl)
    payment_active = models.AdminInfo.objects.filter().first().payment_active
    if not payment_active:
        if request.method == "POST":
            admin = models.AdminInfo.objects.filter().first().phone_number
            user = models.CustomUser.objects.get(id=request.user.id)
            amount = request.POST.get("amount")
            print(amount)
            reference = helper.top_up_ref_generator()
            new_topup_request = models.TopUpRequestt.objects.create(
                user=request.user,
                amount=amount,
                reference=reference,
            )
            new_topup_request.save()

            sms_headers = {
                'Authorization': 'Bearer 1136|LwSl79qyzTZ9kbcf9SpGGl1ThsY0Ujf7tcMxvPze',
                'Content-Type': 'application/json'
            }

            sms_url = 'https://webapp.usmsgh.com/api/sms/send'
            sms_message = f"A top up request has been placed.\nGHS{amount} for {user}.\nReference: {reference}"

            sms_body = {
                'recipient': f"233{admin}",
                'sender_id': 'Geosams',
                'message': sms_message
            }
            # response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
            # print(response.text)
            new_profit_instance = models.ProfitInstance.objects.create(
                selling_price_total=amount,
                channel="Wallet Topup",  # Set your channel here based on user status
            )
            new_profit_instance.save()
            messages.success(request,
                             f"Your Request has been sent successfully. Make payment now")
            return redirect("request_successful", reference)
    db_user_id = request.user.id
    user_email = request.user.email
    reference = helper.ref_generator()
    payment_active = models.AdminInfo.objects.filter().first().payment_active
    context = {'id': db_user_id, "ref": reference, "email": user_email, "payment_active": payment_active}
    return render(request, "layouts/topup-info.html", context=context)


@login_required(login_url='login')
def request_successful(request, reference):
    admin = models.AdminInfo.objects.filter().first()
    context = {
        "name": admin.name,
        "number": f"0{admin.momo_numberr}",
        "channel": admin.payment_channell,
        "reference": reference
    }
    return render(request, "layouts/services/request_successful.html", context=context)


@login_required(login_url='login')
def topup_list(request):
    if request.user.is_superuser:
        topup_requests = models.TopUpRequestt.objects.all().order_by('date').reverse()[:200]
        context = {
            'requests': topup_requests,
        }
        return render(request, "layouts/services/topup_list.html", context=context)
    else:
        messages.error(request, "Access Denied")
        return redirect('home')


@ratelimit(key='ip', rate='10/m')
@login_required(login_url='login')
def credit_user_from_list(request, reference):
    if request.user.is_superuser:
        crediting = models.TopUpRequestt.objects.filter(reference=reference).first()
        if crediting.status:
            return redirect('topup_list')
        user = crediting.user
        custom_user = models.CustomUser.objects.get(username=user.username)
        amount = crediting.amount
        print(user)
        print(user.phone)
        print(amount)
        custom_user.wallet += amount
        custom_user.save()
        print(custom_user.wallet)
        new_wallet_transaction = models.WalletTransaction.objects.create(
            user=custom_user,
            transaction_type="Credit",
            transaction_amount=float(amount),
            transaction_use="Top up",
            new_balance=custom_user.wallet
        )
        new_wallet_transaction.save()
        crediting.status = True
        crediting.credited_at = datetime.now()
        crediting.save()
        sms_headers = {
            'Authorization': 'Bearer 1334|wroIm5YnQD6hlZzd8POtLDXxl4vQodCZNorATYGX',
            'Content-Type': 'application/json'
        }

        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
        sms_message = f"Hello,\nYour wallet has been topped up with GHS{amount}.\nReference: {reference}.\nThank you"

        response1 = requests.get(
            f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=UnBzemdvanJyUGxhTlJzaVVQaHk&to=0{user.phone}&from=GEO_AT&sms={sms_message}")
        print(response1.text)
        messages.success(request, f"{user} has been credited with {amount}")
        return redirect('topup_list')


@csrf_exempt
def hubtel_webhook(request):
    if request.method == 'POST':
        print("hit the webhook")
        try:
            payload = request.body.decode('utf-8')
            print("Hubtel payment Info: ", payload)
            json_payload = json.loads(payload)
            print(json_payload)

            data = json_payload.get('Data')
            print(data)
            reference = data.get('ClientReference')
            print(reference)
            txn_status = data.get('Status')
            txn_description = data.get('Description')
            amount = data.get('Amount')
            print(txn_status, amount)

            if txn_status == 'Success':
                print("success")
                transaction_saved = models.Payment.objects.get(reference=reference, transaction_status="Unfinished")
                transaction_saved.transaction_status = "Paid"
                transaction_saved.payment_description = txn_description
                transaction_saved.amount = amount
                transaction_saved.save()
                transaction_details = transaction_saved.transaction_details
                transaction_channel = transaction_saved.channel
                user = transaction_saved.user
                # receiver = collection_saved['number']
                # bundle_volume = collection_saved['data_volume']
                # name = collection_saved['name']
                # email = collection_saved['email']
                # phone_number = collection_saved['buyer']
                # date_and_time = collection_saved['date_and_time']
                # txn_type = collection_saved['type']
                # user_id = collection_saved['uid']
                print(transaction_details, transaction_channel)

                if transaction_channel == "ishare":
                    offer = transaction_details["offers"]
                    phone_number = transaction_details["phone_number"]

                    if user.status == "User":
                        bundle = models.IshareBundlePrice.objects.get(price=float(offer)).bundle_volume
                    elif user.status == "Agent":
                        bundle = models.AgentIshareBundlePrice.objects.get(price=float(offer)).bundle_volume
                    elif user.status == "Super Agent":
                        bundle = models.SuperAgentIshareBundlePrice.objects.get(price=float(offer)).bundle_volume
                    else:
                        bundle = models.IshareBundlePrice.objects.get(price=float(offer)).bundle_volume
                    new_transaction = models.IShareBundleTransaction.objects.create(
                        user=user,
                        bundle_number=phone_number,
                        offer=f"{bundle}MB",
                        reference=reference,
                        transaction_status="Pending"
                    )
                    print("created")
                    new_transaction.save()

                    print("===========================")
                    print(phone_number)
                    print(bundle)
                    print(user)
                    print(reference)
                    send_bundle_response = helper.send_bundle(user, f"0{phone_number}", bundle, reference)
                    # data = send_bundle_response.json()
                    #
                    # print(data)

                    sms_headers = {
                        'Authorization': 'Bearer 1136|LwSl79qyzTZ9kbcf9SpGGl1ThsY0Ujf7tcMxvPze',
                        'Content-Type': 'application/json'
                    }

                    sms_url = 'https://webapp.usmsgh.com/api/sms/send'

                    if send_bundle_response != "bad response":
                        print("good response")
                        if send_bundle_response["data"]["request_status_code"] == "200" or send_bundle_response[
                            "request_message"] == "Successful":
                            transaction_to_be_updated = models.IShareBundleTransaction.objects.get(
                                reference=reference)
                            print("got here")
                            print(transaction_to_be_updated.transaction_status)
                            transaction_to_be_updated.transaction_status = "Completed"
                            transaction_to_be_updated.save()
                            print(user.phone)
                            print("***********")
                            receiver_message = f"Your bundle purchase has been completed successfully. {bundle}MB has been credited to you by {user.phone}.\nReference: {reference}\n"
                            sms_message = f"Hello @{user.username}. Your bundle purchase has been completed successfully. {bundle}MB has been credited to {phone_number}.\nReference: {reference}\n"

                            response1 = requests.get(
                                f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=UnBzemdvanJyUGxhTlJzaVVQaHk&to=0{user.phone}&from=GEO_AT&sms={sms_message}")
                            print(response1.text)

                            response2 = requests.get(
                                f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=UnBzemdvanJyUGxhTlJzaVVQaHk&to={phone_number}&from=GEO_AT&sms={receiver_message}")
                            print(response2.text)
                            return JsonResponse({'status': 'Transaction Completed Successfully'}, status=200)
                        else:
                            transaction_to_be_updated = models.IShareBundleTransaction.objects.get(
                                reference=reference)
                            transaction_to_be_updated.transaction_status = "Failed"
                            new_transaction.save()
                            sms_message = f"Hello @{user.username}. Something went wrong with your transaction. Contact us for enquiries.\nBundle: {bundle}MB\nPhone Number: {phone_number}.\nReference: {reference}\n"

                            sms_body = {
                                'recipient': f"233{user.phone}",
                                'sender_id': 'Data4All',
                                'message': sms_message
                            }
                            return JsonResponse({'status': 'Something went wrong'}, status=500)
                    else:
                        transaction_to_be_updated = models.IShareBundleTransaction.objects.get(
                            reference=reference)
                        transaction_to_be_updated.transaction_status = "Failed"
                        new_transaction.save()
                        sms_message = f"Hello @{user.username}. Something went wrong with your transaction. Contact us for enquiries.\nBundle: {bundle}MB\nPhone Number: {phone_number}.\nReference: {reference}\n"

                        sms_body = {
                            'recipient': f'233{user.phone}',
                            'sender_id': 'Data4All',
                            'message': sms_message
                        }

                        # response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
                        #
                        # print(response.text)
                        return JsonResponse({'status': 'Something went wrong', 'icon': 'error'})
                elif transaction_channel == "mtn":
                    offer = transaction_details["offers"]
                    phone_number = transaction_details["phone_number"]

                    if user.status == "User":
                        bundle = models.MTNBundlePrice.objects.get(price=float(offer)).bundle_volume
                    elif user.status == "Agent":
                        bundle = models.AgentMTNBundlePrice.objects.get(price=float(offer)).bundle_volume
                    elif user.status == "Super Agent":
                        bundle = models.SuperAgentMTNBundlePrice.objects.get(price=float(offer)).bundle_volume
                    else:
                        bundle = models.MTNBundlePrice.objects.get(price=float(offer)).bundle_volume

                    print(phone_number)
                    new_mtn_transaction = models.MTNTransaction.objects.create(
                        user=user,
                        bundle_number=phone_number,
                        offer=f"{bundle}MB",
                        reference=reference,
                    )
                    new_mtn_transaction.save()
                    return JsonResponse({'status': "Your transaction will be completed shortly"}, status=200)
                elif transaction_channel == "bigtime":
                    offer = transaction_details["offers"]
                    phone_number = transaction_details["phone_number"]
                    if user.status == "User":
                        bundle = models.BigTimeBundlePrice.objects.get(price=float(offer)).bundle_volume
                    elif user.status == "Agent":
                        bundle = models.AgentBigTimeBundlePrice.objects.get(price=float(offer)).bundle_volume
                    elif user.status == "Super Agent":
                        bundle = models.SuperAgentBigTimeBundlePrice.objects.get(price=float(offer)).bundle_volume
                    else:
                        bundle = models.SuperAgentBigTimeBundlePrice.objects.get(price=float(offer)).bundle_volume
                    print(phone_number)
                    new_mtn_transaction = models.BigTimeTransaction.objects.create(
                        user=user,
                        bundle_number=phone_number,
                        offer=f"{bundle}MB",
                        reference=reference,
                    )
                    new_mtn_transaction.save()
                    return JsonResponse({'status': "Your transaction will be completed shortly"}, status=200)
                elif transaction_channel == "afa":
                    name = transaction_details["name"]
                    phone_number = transaction_details["phone"]
                    gh_card_number = transaction_details["card"]
                    occupation = transaction_details["occupation"]
                    date_of_birth = transaction_details["date_of_birth"]

                    new_afa_reg = models.AFARegistration.objects.create(
                        user=user,
                        phone_number=phone_number,
                        gh_card_number=gh_card_number,
                        name=name,
                        occupation=occupation,
                        reference=reference,
                        date_of_birth=date_of_birth
                    )
                    new_afa_reg.save()
                    return JsonResponse({'status': "Your transaction will be completed shortly"}, status=200)
                elif transaction_channel == "topup":
                    amount = amount
                    amount = round(amount - (amount * 0.01))

                    user.wallet += float(amount)
                    user.save()

                    new_topup = models.TopUpRequestt.objects.create(
                        user=user,
                        reference=reference,
                        amount=amount,
                        status=True,
                    )
                    new_topup.save()

                    new_wallet_transaction = models.WalletTransaction.objects.create(
                        user=user,
                        transaction_type="Credit",
                        transaction_amount=float(amount),
                        transaction_use="Top up (Hubtel)",
                        new_balance=user.wallet
                    )
                    new_wallet_transaction.save()
                    return JsonResponse({'status': "Wallet Credited"}, status=200)
                elif transaction_channel == "commerce":
                    name = transaction_details["name"]
                    email = transaction_details["email"]
                    phone = transaction_details["phone"]
                    address = transaction_details["address"]
                    city = transaction_details["city"]
                    region = transaction_details["region"]
                    message = transaction_details["message"]

                    new_order_items = models.Cart.objects.filter(user=user)
                    cart = models.Cart.objects.filter(user=user)
                    cart_total_price = 0
                    for item in cart:
                        cart_total_price += item.product.selling_price * item.product_qty
                    print(cart_total_price)

                    order_instance = models.Order.objects.create(
                        user=user,
                        full_name=name,
                        email=email,
                        phone=phone,
                        address=address,
                        city=city,
                        region=region,
                        total_price=cart_total_price,
                        payment_mode="Hubtel",
                        message=message,
                        tracking_number=reference,
                    )
                    order_instance.save()

                    for item in new_order_items:
                        models.OrderItem.objects.create(
                            order=order_instance,
                            product=item.product,
                            tracking_number=order_instance.tracking_number,
                            price=item.product.selling_price,
                            quantity=item.product_qty
                        )
                        order_product = models.Product.objects.filter(id=item.product_id).first()
                        order_product.quantity -= item.product_qty
                        order_product.save()

                    models.Cart.objects.filter(user=user).delete()

                    sms_message = f"Order Placed Successfully\nYour order with order number {order_instance.tracking_number} has been received and is being processed.\nYou will receive a message when your order is Out for Delivery.\nThank you for shopping with GeoSams"

                    try:
                        response1 = requests.get(
                            f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=UnBzemdvanJyUGxhTlJzaVVQaHk&to=0{order_instance.phone}&from=GEO_AT&sms={sms_message}")
                        print(response1.text)
                    except:
                        print("Could not send sms message")
                    return JsonResponse({'message': "Order Placed"}, status=200)
                elif transaction_channel == "agent":
                    amount = transaction_details["amount"]

                    new_registration = models.AgentRegistration.objects.create(
                        amount=amount,
                        user=user
                    )
                    new_registration.save()

                    user.status = "Agent"
                    user.save()
                    return JsonResponse({'message': 'Transaction Successful'}, status=200)
                else:
                    print("no type found")
                    return JsonResponse({'message': "No Type Found"}, status=500)
            else:
                print("Transaction was not Successful")
                return JsonResponse({'message': 'Transaction Failed'}, status=200)
        except Exception as e:
            print("Error Processing hubtel webhook:", str(e))
            return JsonResponse({'status': 'error'}, status=500)
    else:
        print("not post")
        return JsonResponse({'message': 'Not Found'}, status=404)


def populate_custom_users_from_excel(request):
    # Read the Excel file using pandas
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['file']

            # Process the uploaded Excel file
            df = pd.read_excel(excel_file)
            counter = 0
            # Iterate through rows to create CustomUser instances
            for index, row in df.iterrows():
                print(counter)
                # Create a CustomUser instance for each row
                custom_user = CustomUser.objects.create(
                    first_name=row['first_name'],
                    last_name=row['last_name'],
                    username=str(row['username']),
                    email=row['email'],
                    phone=row['phone'],
                    wallet=float(row['wallet']),
                    status=str(row['status']),
                    password1=row['password1'],
                    password2=row['password2'],
                    is_superuser=row['is_superuser'],
                    is_staff=row['is_staff'],
                    is_active=row['is_active'],
                    password=row['password']
                )

                custom_user.save()

                # group_names = row['groups'].split(',')  # Assuming groups are comma-separated
                # groups = Group.objects.filter(name__in=group_names)
                # custom_user.groups.set(groups)
                #
                # if row['user_permissions']:
                #     permission_ids = [int(pid) for pid in row['user_permissions'].split(',')]
                #     permissions = Permission.objects.filter(id__in=permission_ids)
                #     custom_user.user_permissions.set(permissions)
                print("killed")
                counter = counter + 1
            messages.success(request, 'All done')
    else:
        form = UploadFileForm()
    return render(request, 'layouts/import_users.html', {'form': form})


def delete_custom_users(request):
    CustomUser.objects.all().delete()
    return HttpResponseRedirect('Done')


@csrf_exempt
def initiate_mtn_transaction(request):
    if request.method == "POST":
        print(request.POST)
        receiver = request.POST.get("receiver")
        bundle_volume = request.POST.get("bundle_volume")
        reference = request.POST.get("reference")

        if not receiver or not bundle_volume or not reference:
            return JsonResponse({'message': 'Missing Parameters'}, status=400)

        print(receiver)
        print(bundle_volume)

        api_key = request.headers.get("api-key")
        if not api_key:
            return JsonResponse({'message': 'No Api Key Found'}, status=401)

        try:
            api_user = models.MTNAPIUsers.objects.get(key=api_key)
            print(api_user.key)
        except Exception as e:
            print(e)
            return JsonResponse({'message': 'Invalid API Key'}, status=401)

        try:
            bundle_price = models.APIMTNBundlePrice.objects.get(bundle_volume=bundle_volume).price
        except models.APIMTNBundlePrice.DoesNotExist:
            return JsonResponse({'message': 'Bundle Volume not found'}, status=400)

        try:
            api_user = models.MTNAPIUsers.objects.get(key=api_key)
            wallet_balance = api_user.wallet_balance
            print(wallet_balance)

            if wallet_balance < float(bundle_price):
                return JsonResponse({'message': 'Insufficient balance'}, status=400)

            api_user.wallet_balance -= float(bundle_price)
            api_user.save()

            print(api_user.user)

            new_history = models.ApiWalletTransaction.objects.create(
                user=api_user.user,
                transaction_type="Debit",
                transaction_amount=float(bundle_price),
                new_balance=float(api_user.wallet_balance),
                transaction_channel="MTN"
            )
            new_history.save()

            new_mtn_transaction = models.MTNTransaction.objects.create(
                user=api_user.user,
                bundle_number=receiver,
                offer=f"{bundle_volume}MB",
                reference=reference,
            )
            new_mtn_transaction.save()

            print("saved")

            new_api_history = models.APIUsersHistory.objects.create(
                mtn_transaction=new_mtn_transaction,
                api_user=api_user
            )

            new_api_history.save()

            print("saved")

            return JsonResponse({'status': "Your transaction will be completed shortly"}, status=200)
        except models.MTNAPIUsers.DoesNotExist:
            return JsonResponse(data={"message": "Invalid API Key"}, status=401)


@csrf_exempt
def initiate_telecel_transaction(request):
    if request.method == "POST":
        print(request.POST)
        receiver = request.POST.get("receiver")
        bundle_volume = request.POST.get("bundle_volume")
        reference = request.POST.get("reference")

        if not receiver or not bundle_volume or not reference:
            return JsonResponse({'message': 'Missing Parameters'}, status=400)

        print(receiver)
        print(bundle_volume)

        api_key = request.headers.get("api-key")
        if not api_key:
            return JsonResponse({'message': 'No Api Key Found'}, status=401)

        try:
            api_user = models.MTNAPIUsers.objects.get(key=api_key)
            print(api_user.key)
        except Exception as e:
            print(e)
            return JsonResponse({'message': 'Invalid API Key'}, status=401)

        try:
            bundle_price = models.APITelecelBundlePrice.objects.get(bundle_volume=bundle_volume).price
        except models.APITelecelBundlePrice.DoesNotExist:
            return JsonResponse({'message': 'Bundle Volume not found'}, status=400)

        try:
            api_user = models.MTNAPIUsers.objects.get(key=api_key)
            wallet_balance = api_user.wallet_balance
            print(wallet_balance)

            if wallet_balance < float(bundle_price):
                return JsonResponse({'message': 'Insufficient balance'}, status=400)

            api_user.wallet_balance -= float(bundle_price)
            api_user.save()

            print(api_user.user)

            new_history = models.ApiWalletTransaction.objects.create(
                user=api_user.user,
                transaction_type="Debit",
                transaction_amount=float(bundle_price),
                new_balance=float(api_user.wallet_balance),
                transaction_channel="Telecel"
            )
            new_history.save()

            new_telecel_transaction = models.VodafoneTransaction.objects.create(
                user=api_user.user,
                bundle_number=receiver,
                offer=f"{bundle_volume}MB",
                reference=reference,
            )
            new_telecel_transaction.save()

            print("saved")

            new_api_history = models.TelecelAPIUsersHistory.objects.create(
                telecel_transaction=new_telecel_transaction,
                api_user=api_user
            )

            new_api_history.save()

            print("saved")

            return JsonResponse({'status': "Your transaction will be completed shortly"}, status=200)
        except models.MTNAPIUsers.DoesNotExist:
            return JsonResponse(data={"message": "Invalid API Key"}, status=401)


from django.utils.encoding import force_bytes
from django.contrib.auth.tokens import default_token_generator


@ratelimit(key='ip', rate='3/m')
def password_reset_request(request):
    if request.method == "POST":
        password_reset_form = PasswordResetForm(request.POST)
        if password_reset_form.is_valid():
            data = password_reset_form.cleaned_data['email']
            user = models.CustomUser.objects.filter(email=data).first()
            current_user = user
            if user:
                subject = "Password Reset Requested"
                email_template_name = "password/password_reset_message.txt"
                c = {
                    "name": user.first_name,
                    "email": user.email,
                    'domain': 'www.geosams.com',
                    'site_name': 'GeoSams',
                    "uid": urlsafe_base64_encode(force_bytes(user.pk)),
                    "user": user,
                    'token': default_token_generator.make_token(user),
                    'protocol': 'https',
                }
                email = render_to_string(email_template_name, c)

                # sms_headers = {
                #     'Authorization': 'Bearer 1317|sCtbw8U97Nwg10hVbZLBPXiJ8AUby7dyozZMjJpU',
                #     'Content-Type': 'application/json'
                # }
                #
                # sms_url = 'https://webapp.usmsgh.com/api/sms/send'
                #
                # sms_body = {
                #     'recipient': f"233{user.phone}",
                #     'sender_id': 'GH DATA',
                #     'message': email
                # }
                # response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
                # print(response.text)
                response1 = requests.get(
                    f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=UnBzemdvanJyUGxhTlJzaVVQaHk&to=0{user.phone}&from=GEO_AT&sms={email}")
                print(response1.text)

                return redirect("/password_reset/done/")
    password_reset_form = PasswordResetForm()
    return render(request=request, template_name="password/password_reset.html",
                  context={"password_reset_form": password_reset_form})


@csrf_exempt
def paystack_webhook(request):
    if request.method == "POST":
        paystack_secret_key = config("PAYSTACK_SECRET_KEY")
        # print(paystack_secret_key)
        payload = json.loads(request.body)

        paystack_signature = request.headers.get("X-Paystack-Signature")

        if not paystack_secret_key or not paystack_signature:
            return HttpResponse(status=400)

        computed_signature = hmac.new(
            paystack_secret_key.encode('utf-8'),
            request.body,
            hashlib.sha512
        ).hexdigest()

        if computed_signature == paystack_signature:
            print("yes")
            print(payload.get('data'))
            r_data = payload.get('data')
            print(r_data.get('metadata'))
            print(payload.get('event'))
            if payload.get('event') == 'charge.success':
                metadata = r_data.get('metadata')
                receiver = metadata.get('receiver')
                db_id = metadata.get('db_id')
                referer = metadata.get('referrer')
                print(referer)
                if referer != "https://www.geosams.com/topup-info":
                    print("invalid referrer")
                    return HttpResponse(status=200)
                print(db_id)
                # offer = metadata.get('offer')
                user = models.CustomUser.objects.get(id=int(db_id))
                print(user)
                channel = metadata.get('channel')
                real_amount = metadata.get('real_amount')
                print(real_amount)
                paid_amount = r_data.get('amount')
                slashed_amount = float(paid_amount) / 100
                reference = r_data.get('reference')

                rounded_real_amount = round(float(real_amount))
                rounded_paid_amount = float(slashed_amount)
                deducted_paid_amount = slashed_amount - ((1 / 100) * rounded_paid_amount)

                print(f"reeeeeeeaaaaaaaaal amount: {rounded_real_amount}")
                print(f"paaaaaaaaaaaaaiiddd amount: {rounded_paid_amount}")

                is_within_range = (rounded_real_amount - 20) <= rounded_paid_amount <= (rounded_real_amount + 20)

                if not is_within_range:
                    # sms_headers = {
                    #     'Authorization': 'Bearer 1317|sCtbw8U97Nwg10hVbZLBPXiJ8AUby7dyozZMjJpU',
                    #     'Content-Type': 'application/json'
                    # }
                    #
                    # sms_url = 'https://webapp.usmsgh.com/api/sms/send'
                    # sms_message = f"Malicious attempt on webhook. Real amount: {rounded_real_amount} | Paid amount: {rounded_paid_amount}. Referrer: {reference}"
                    #
                    # sms_body = {
                    #     'recipient': "233242442147",
                    #     'sender_id': 'GH DATA',
                    #     'message': sms_message
                    # }
                    # try:
                    #     response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
                    #     print(response.text)
                    # except Exception as e:
                    #     print(e)

                    print("not within range")
                    return HttpResponse(200)

                if channel == "topup":
                    print("topupppppppppppp")
                    topup_amount = metadata.get('real_amount')

                    if models.TopUpRequestt.objects.filter(user=user, reference=reference).exists():
                        return HttpResponse(status=200)

                    new_payment = models.Payment.objects.create(
                        user=user,
                        reference=reference,
                        amount=paid_amount,
                        transaction_date=datetime.datetime.now(),
                        transaction_status="Completed"
                    )
                    new_payment.save()
                    print(user.wallet)
                    user.wallet += float(topup_amount)
                    user.save()
                    print(user.wallet)

                    if models.TopUpRequestt.objects.filter(user=user, reference=reference, status=True).exists():
                        return HttpResponse(status=200)

                    new_topup = models.TopUpRequestt.objects.create(
                        user=user,
                        reference=reference,
                        amount=topup_amount,
                        status=True,
                    )
                    new_topup.save()

                    new_profit_instance = models.ProfitInstance.objects.create(
                        selling_price_total=topup_amount,
                        channel="Wallet Topup",  # Set your channel here based on user status
                    )
                    new_profit_instance.save()

                    sms_headers = {
                        'Authorization': 'Bearer 1317|sCtbw8U97Nwg10hVbZLBPXiJ8AUby7dyozZMjJpU',
                        'Content-Type': 'application/json'
                    }

                    sms_url = 'https://webapp.usmsgh.com/api/sms/send'
                    sms_message = f"Your Geosams wallet has been credited with GHS{topup_amount}.\nReference: {reference}\n"

                    sms_body = {
                        'recipient': f"233{user.phone}",
                        'sender_id': 'GH DATA',
                        'message': sms_message
                    }
                    try:
                        response1 = requests.get(
                            f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=UnBzemdvanJyUGxhTlJzaVVQaHk&to=0{user.phone}&from=GEO_AT&sms={sms_message}")
                        print(response1.text)
                        return HttpResponse(status=200)
                    except:
                        return HttpResponse(status=200)
                else:
                    return HttpResponse(status=200)
            else:
                return HttpResponse(status=200)
        else:
            return HttpResponse(status=401)
    else:
        return HttpResponse(status=200)


def cancel_mtn_transaction(request, pk):
    user = models.CustomUser.objects.get(id=request.user.id)
    try:
        transaction_to_be_canceled = models.MTNTransaction.objects.filter(id=pk, user=user,
                                                                          transaction_status="Pending").first()
    except Exception as e:
        print(e)
        messages.info(request, "Could not cancel transaction")
        return redirect('mtn-history')

    try:
        amount_to_refund = transaction_to_be_canceled.amount
        transaction_to_be_canceled.transaction_status = "Canceled"
        transaction_to_be_canceled.save()
        try:
            user.wallet += float(amount_to_refund)
            user.save()
        except Exception as e:
            print(e)
            user.wallet += int(amount_to_refund)
            user.save()

        new_wallet_transaction = models.WalletTransaction.objects.create(
            user=user,
            transaction_type="Credit",
            transaction_amount=float(amount_to_refund),
            transaction_use="Refund(MTN)",
            new_balance=user.wallet
        )
        new_wallet_transaction.save()

        new_profit_instance = models.ProfitInstance.objects.create(
            selling_price_total=amount_to_refund,
            channel="Refunds",  # Set your channel here based on user status
        )
        new_profit_instance.save()

    except Exception as e:
        print(e)
        messages.info(request, "Unable to cancel transaction")
        return redirect('mtn-history')

    messages.success(request, "Transaction has been cancelled and money refunded into wallet")

    return redirect('mtn-history')


def admin_cancel_mtn_transaction(request, pk):
    if request.user.is_superuser and request.user.is_staff:
        transaction_to_be_canceled = models.MTNTransaction.objects.get(id=pk)
        user = transaction_to_be_canceled.user
        print(user)
        print(user.wallet)
        try:
            amount_to_refund = transaction_to_be_canceled.amount
            transaction_to_be_canceled.transaction_status = "Canceled"
            transaction_to_be_canceled.save()
            try:
                user.wallet += float(amount_to_refund)
                user.save()
            except Exception as e:
                print(e)
                user.wallet += int(amount_to_refund)
                user.save()

            print(user.wallet)

            new_wallet_transaction = models.WalletTransaction.objects.create(
                user=user,
                transaction_type="Credit",
                transaction_amount=float(amount_to_refund),
                transaction_use="Refund(MTN)",
                new_balance=user.wallet
            )
            new_wallet_transaction.save()

            new_profit_instance = models.ProfitInstance.objects.create(
                selling_price_total=amount_to_refund,
                channel="Refunds",  # Set your channel here based on user status
            )
            new_profit_instance.save()

        except Exception as e:
            print(e)
            messages.info(request, "Unable to cancel transaction")
            return redirect('mtn-history')

        messages.success(request, "Transaction has been cancelled and money refunded into wallet")

        return redirect('mtn_admin', status="Pending")
    else:
        return redirect('home')


@login_required(login_url='login')
def profit_home(request):
    if request.user.is_superuser:
        return render(request, 'layouts/profit_page.html')
    else:
        messages.error(request, "Access Denied")
        return redirect('home')


def channel_profit(request, channel):
    profit_instances = models.ProfitInstance.objects.filter(channel=channel)

    # Calculate profit breakdown by month
    today = timezone.now()
    this_year = today.year
    monthly_data = []
    total_profit = 0
    total_selling_price = 0
    total_purchase_price = 0

    for month in range(1, 13):  # Loop through each month of the year
        month_name = calendar.month_name[month]
        month_instances = profit_instances.filter(date_and_time__year=this_year, date_and_time__month=month)
        month_profit = month_instances.aggregate(total_profit=Sum('profit'))['total_profit'] or 0
        month_selling_price = month_instances.aggregate(total_selling_price=Sum('selling_price_total'))[
                                  'total_selling_price'] or 0
        # month_purchase_price = month_instances.aggregate(total_purchase_price=Sum('purchase_price_total'))[
        #                            'total_purchase_price'] or 0

        monthly_data.append({
            'month': month_name,
            'profit': month_profit,
            'selling_price': month_selling_price,
        })

        total_profit += month_profit
        total_selling_price += month_selling_price
        # total_purchase_price += month_purchase_price

    context = {
        'monthly_data': monthly_data,
        'total_profit': total_profit,
        'total_selling_price': total_selling_price,
        'total_purchase_price': total_purchase_price,
        'channel': channel,
    }

    return render(request, 'layouts/services/profit.html', context)
